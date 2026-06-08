# -*- coding: utf-8 -*-
"""Выгрузки для Яндекс.Директа: CSV семантики, тексты объявлений, прогноз по радиусу, Коммандер-XLSX."""
import json, csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE="/Users/zif91/Programs_py/КОсмос/"
DL="/Users/zif91/Downloads/"
fc=json.load(open("/tmp/forecast.json"))
CAMP="Космос Детейлинг — Поиск СПб"

GROUP_ORDER=["Детейлинг общий","Детейлинг · Васильевский остров","Детейлинг-мойка","Оклейка / PPF",
"Полировка кузова","Керамика / защита","Тонировка","Оптика / фары","Химчистка и кожа","Предпродажная подготовка"]
EFF_BASE=0.35  # IS 50% × overlap 0.70 (как в медиаплане)
for r in fc:
    pc=r.get("prem_clicks",0) or 0
    r["rc_base"]=pc*EFF_BASE
    r["bud_min"]=r["rc_base"]*(r.get("prem_min",0) or 0)
    r["bud_max"]=r["rc_base"]*(r.get("prem_max",0) or 0)

groups={g:[r for r in fc if r["group"]==g] for g in GROUP_ORDER}
def gsum(rows,k): return sum((r.get(k,0) or 0) for r in rows)

# рекомендованная стартовая ставка = вход в премиум (PremiumMin), но не выше разумного потолка
def bid(r):
    pm=r.get("prem_min",0) or 0
    if pm<=0: return 30
    return int(round(min(pm, (r.get("prem_max",0) or pm))))

# ---------- 1. CSV СЕМАНТИКИ ----------
def w_csv(path,header,rows):
    with open(path,"w",encoding="utf-8-sig",newline="") as f:
        w=csv.writer(f,delimiter=";"); w.writerow(header); w.writerows(rows)
sem_rows=[]
for g in GROUP_ORDER:
    for r in sorted(groups[g],key=lambda x:-(x.get("prem_clicks",0) or 0)):
        sem_rows.append([CAMP,g,f'"{r["kw"]}"',"фразовое","Санкт-Петербург",bid(r),r["landing"]])
w_csv(BASE+"Космос_Директ_семантика.csv",
      ["Кампания","Группа объявлений","Ключевая фраза","Тип соответствия","Регион","Реком. ставка CPC, ₽","Посадочная"],sem_rows)

# фразы по группам (для быстрой вставки)
with open(BASE+"Космос_фразы_по_группам.csv","w",encoding="utf-8-sig",newline="") as f:
    w=csv.writer(f,delimiter=";"); w.writerow(["Группа","Фраза (фразовое соответствие)"])
    for g in GROUP_ORDER:
        for r in sorted(groups[g],key=lambda x:-(x.get("prem_clicks",0) or 0)):
            w.writerow([g,f'"{r["kw"]}"'])

# ---------- 2. ТЕКСТЫ ОБЪЯВЛЕНИЙ ----------
# (Заголовок1<=56, Заголовок2<=30, Текст<=81, уточнение<=25, быстрая ссылка<=30, путь<=20)
DISP={"Детейлинг общий":"detailing","Детейлинг · Васильевский остров":"vo","Детейлинг-мойка":"moyka",
"Оклейка / PPF":"ppf","Полировка кузова":"polirovka","Керамика / защита":"keramika","Тонировка":"tonirovka",
"Оптика / фары":"fary","Химчистка и кожа":"himchistka","Предпродажная подготовка":"predprodazha"}
SL_COMMON=[("Услуги и цены","{land}"),("Наши работы","index.html#raboty"),("Пакеты","index.html#pakety"),("Контакты","kontakty.html")]
CALLOUTS=["Фиксированная смета","Фото- и видео-отчёт","Гарантия на работы","Отдельные чистые боксы","Васильевский остров","Запись онлайн"]
ADS={
"Детейлинг общий":[
 ("Детейлинг на Васильевском острове","Фикс. смета и фотоотчёт","Премиум-детейлинг в отдельных боксах. Смета до работ, гарантия. Запишитесь!"),
 ("Космос Детейлинг — студия в СПб","Это не мойка, а детейлинг","Полировка, керамика, химчистка, защита. Фото/видео-отчёт. Кораблестроителей 14."),
 ("Детейлинг-студия рядом, В.О.","Запись на удобное время","Чистые боксы, мастера, гарантия результата. Честная смета. Звоните!"),
],
"Детейлинг · Васильевский остров":[
 ("Детейлинг на Васильевском острове","Рядом, без очередей","Студия на Кораблестроителей 14. Фикс. смета, фотоотчёт, гарантия. Запись!"),
 ("Детейлинг рядом — Намыв, В.О.","Премиум-уход за авто","Полировка, керамика, химчистка. Отчёт в мессенджер. Запишитесь онлайн!"),
],
"Детейлинг-мойка":[
 ("Детейлинг-мойка авто в СПб","Аккуратно, в две фазы","Бесконтактная мойка + уход за кузовом и салоном. Васильевский остров. Запись!"),
 ("Детейлинг-мойка на Васильевском","Бережно к ЛКП","Двухфазная мойка, сушка, защита. Чистый бокс, не поток. Звоните!"),
],
"Оклейка / PPF":[
 ("Оклейка авто плёнкой в СПб","PPF и бронирование","Антигравийная защита кузова. Фикс. смета, гарантия, чистый бокс. Рассчитать!"),
 ("Бронирование авто плёнкой PPF","Защита от сколов","Полиуретан премиум, оклейка по лекалам. Сохраним ЛКП. Кораблестроителей 14."),
 ("Оклейка авто полиуретаном, В.О.","Смета до начала работ","Защитим капот, бампер, зоны риска. Гарантия и фотоотчёт. Запишитесь!"),
],
"Полировка кузова":[
 ("Полировка кузова авто в СПб","Уберём царапины","Восстановительная полировка, глубокий блеск. Фикс. смета, гарантия. Запись!"),
 ("Полировка авто на Васильевском","Замер толщины ЛКП","Безопасная полировка, фото до/после. Отдельные боксы. Рассчитать стоимость!"),
 ("Глубокая полировка кузова, СПб","Блеск как новый","Уберём голограммы и затёртости. Гарантия результата. Кораблестроителей 14."),
],
"Керамика / защита":[
 ("Керамика на авто в СПб","Защита и глубокий блеск","Керамическое покрытие в несколько слоёв. Гарантия, фотоотчёт. В.О."),
 ("Керамическое покрытие авто","Гидрофоб и стойкость","Полировка + керамика премиум. Фикс. смета до работ. Запишитесь онлайн!"),
],
"Тонировка":[
 ("Тонировка авто в СПб","По ГОСТ и атермальная","Плёнки премиум, ровно, без пузырей. Гарантия. Кораблестроителей 14. Запись!"),
 ("Тонировка стёкол на Васильевском","Атермалка — меньше жары","Подберём допустимую тонировку. Аккуратный монтаж, гарантия. Звоните!"),
],
"Оптика / фары":[
 ("Бронирование и полировка фар","Защита оптики","Вернём прозрачность фар, защитим плёнкой. Фикс. смета, гарантия. СПб, В.О."),
 ("Полировка фар в СПб","Прозрачность как новые","Шлифовка + защита фар и лобового. Запишитесь на Васильевском острове!"),
],
"Химчистка и кожа":[
 ("Химчистка салона авто в СПб","Сухой салон, без запаха","Глубокая химчистка + озонация. Гарантия сухости. Васильевский остров. Запись!"),
 ("Химчистка салона на Васильевском","Фикс. смета, фотоотчёт","Безопасная химия, реставрация кожи. Выдаём по проверке. Рассчитать!"),
 ("Детейлинг-химчистка салона, СПб","Уберём пятна и запах","Полная химчистка, озонация, уход за кожей. Гарантия. Кораблестроителей 14."),
],
"Предпродажная подготовка":[
 ("Предпродажная подготовка авто","Продайте дороже","Мойка, полировка, химчистка, оптика. Снизим поводы для торга. СПб, В.О."),
 ("Предпродажная подготовка, СПб","Товарный вид авто","Комплекс детейлинга к продаже. Фикс. смета, фотоотчёт. Запишитесь!"),
],
}
LIM={"t1":56,"t2":30,"tx":81,"co":25,"sl":30,"path":20}
warn=[]
def chk(tag,s,lim):
    if len(s)>lim: warn.append(f"{tag}: {len(s)}>{lim} :: {s}")
for g,ads in ADS.items():
    for t1,t2,tx in ads:
        chk("Заг1",t1,56);chk("Заг2",t2,30);chk("Текст",tx,81)
for c in CALLOUTS: chk("Уточн",c,25)

# ---------- РАСЧЁТ РАДИУСА ----------
RAD_LO,RAD_HI=0.10,0.15   # доля городского спроса в радиусе ~5-7 км вокруг Кораблестроителей 14
tot_rc=gsum(fc,"rc_base"); tot_bmin=gsum(fc,"bud_min"); tot_bmax=gsum(fc,"bud_max"); tot_shows=gsum(fc,"d_shows")

# ---------- ЗАПИСЬ ДОП. ЛИСТОВ В ОСНОВНОЙ XLSX ----------
MP=BASE+"Космос_Детейлинг_Яндекс_Директ_медиаплан.xlsx"
wb=openpyxl.load_workbook(MP)
NAVY="0B1F3A";GOLD="C8A24B";GOLDL="F3E9CE";GREY="F2F4F7";WHITE="FFFFFF";GREEN="E7F6EC";BLUE="EAF1FB";LINE="D9DEE7"
def F(sz=11,b=False,color="1A1A1A"): return Font(name="Calibri",size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color=LINE); border=Border(left=thin,right=thin,top=thin,bottom=thin)
def setw(ws,ws_widths):
    for i,w in enumerate(ws_widths,1): ws.column_dimensions[get_column_letter(i)].width=w
def C(ws,r,c,v,font=None,fillc=None,al="left",bord=False,wrap=False,fmt=None):
    x=ws.cell(r,c,v)
    if font:x.font=font
    if fillc:x.fill=fill(fillc)
    x.alignment=Alignment(horizontal=al,vertical="center",wrap_text=wrap)
    if bord:x.border=border
    if fmt:x.number_format=fmt
    return x

# remove if re-run
for nm in ["7. Тексты объявлений","8. Прогноз по радиусу"]:
    if nm in wb.sheetnames: del wb[nm]

# --- Лист 7: Тексты объявлений ---
ws=wb.create_sheet("7. Тексты объявлений"); ws.sheet_view.showGridLines=False
setw(ws,[4,26,40,30,58,14,40,40])
ws.merge_cells("A1:H1"); C(ws,1,1,"ТЕКСТЫ ОБЪЯВЛЕНИЙ ПО ГРУППАМ (лимиты Директа соблюдены)",F(13,True,WHITE),NAVY); ws.row_dimensions[1].height=24
hd=["№","Группа","Заголовок 1 (≤56)","Заголовок 2 (≤30)","Текст (≤81)","Отобр. ссылка","Быстрые ссылки","Уточнения"]
for c,h in enumerate(hd,1): C(ws,2,c,h,F(9,True,WHITE),NAVY,"center",True,wrap=True)
ws.row_dimensions[2].height=30
r=3;n=0
for g in GROUP_ORDER:
    ads=ADS.get(g,[]); land=groups[g][0]["landing"] if groups[g] else "index.html"
    sl=" · ".join(t for t,_ in SL_COMMON)
    co=" · ".join(CALLOUTS[:4])
    for t1,t2,tx in ads:
        n+=1
        vals=[n,g,t1,t2,tx,f"kosmos-detail.ru/{DISP[g]}",sl,co]
        for c,v in enumerate(vals,1):
            C(ws,r,c,v,F(9,c in(3,4)),GREEN if n%2 else WHITE,"left" if c>1 else "center",True,wrap=True)
        ws.row_dimensions[r].height=30; r+=1
C(ws,r+1,2,"Посадочные:",F(9,True,NAVY))
C(ws,r+1,3,"Заголовок 1 содержит ключ группы. Быстрые ссылки и уточнения общие; при желании можно персонализировать под услугу. Пути отобр. ссылки — латиницей.",F(9,False,"555555"),wrap=True)
ws.merge_cells(start_row=r+1,start_column=3,end_row=r+1,end_column=8)

# --- Лист 8: Прогноз по радиусу ---
ws=wb.create_sheet("8. Прогноз по радиусу"); ws.sheet_view.showGridLines=False
setw(ws,[4,34,22,22,22])
ws.merge_cells("A1:E1"); C(ws,1,1,"ПРОГНОЗ С РАДИУСОМ ВОКРУГ КОРАБЛЕСТРОИТЕЛЕЙ 14 (оценка)",F(13,True,WHITE),NAVY); ws.row_dimensions[1].height=24
r=3
note=["Важно: прогноз Яндекс.Директа (CreateNewForecast) считает только по региону (СПб), радиус как гео-настройку API прогноза не принимает.",
"Поэтому радиус оценён масштабированием городского прогноза на долю спроса в зоне ~5–7 км вокруг адреса (Васильевский о., Намыв, о. Декабристов, часть Петроградского/Приморского).",
"Доля принята 10–15% городского поиска. Цена клика (CPC) при радиусе НЕ снижается; локальная релевантность может немного повысить CTR."]
for t in note:
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5); C(ws,r,1,t,F(10,False,"333333"),GREY if r%2 else WHITE,"left",True,wrap=True); ws.row_dimensions[r].height=30; r+=1
r+=1
C(ws,r,2,"Показатель / мес",F(10,True,WHITE),NAVY,"center",True)
C(ws,r,3,"Вся СПб (база)",F(10,True,WHITE),NAVY,"center",True)
C(ws,r,4,"Радиус ~5–7 км (оценка)",F(10,True,WHITE),NAVY,"center",True)
C(ws,r,5,"Доля от города",F(10,True,WHITE),NAVY,"center",True); r+=1
INT="#,##0"
rows=[("Прогноз показов",tot_shows,(tot_shows*RAD_LO,tot_shows*RAD_HI)),
("Реалистичные клики (IS 50%)",tot_rc,(tot_rc*RAD_LO,tot_rc*RAD_HI)),
("Бюджет, ₽ (мин–макс)",None,None)]
C(ws,r,2,"Прогноз показов",F(10),WHITE,"left",True); C(ws,r,3,round(tot_shows),F(10),WHITE,"center",True,fmt=INT); C(ws,r,4,f"{round(tot_shows*RAD_LO):,}–{round(tot_shows*RAD_HI):,}".replace(',',' '),F(10),GOLDL,"center",True); C(ws,r,5,"10–15%",F(10),WHITE,"center",True); r+=1
C(ws,r,2,"Реалистичные клики (IS 50%)",F(10),WHITE,"left",True); C(ws,r,3,round(tot_rc),F(10),WHITE,"center",True,fmt=INT); C(ws,r,4,f"{round(tot_rc*RAD_LO)}–{round(tot_rc*RAD_HI)}",F(10,True),GOLDL,"center",True); C(ws,r,5,"10–15%",F(10),WHITE,"center",True); r+=1
C(ws,r,2,"Бюджет/мес, ₽",F(10),WHITE,"left",True); C(ws,r,3,f"{round(tot_bmin):,}–{round(tot_bmax):,}".replace(',',' '),F(10),WHITE,"center",True); C(ws,r,4,f"{round(tot_bmin*RAD_LO):,}–{round(tot_bmax*RAD_HI):,}".replace(',',' '),F(10,True),GOLDL,"center",True); C(ws,r,5,"10–15%",F(10),WHITE,"center",True); r+=2
rec=["РЕКОМЕНДАЦИЯ: для поиска НЕ ограничивать показы жёстким радиусом — потеряете до 85–90% спроса (запросы «...спб» идут со всего города).",
"Оптимально: таргет на всю СПб + повышающая корректировка ставок на радиус 5–7 км (например, +20–50%) и упор на локальность в текстах («рядом, Васильевский остров»).",
"Жёсткий радиус оправдан только для имиджевых/охватных кампаний в РСЯ или при очень узком бюджете на гиперлокальный старт."]
for t in rec:
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5); C(ws,r,1,t,F(10,False,"222222"),GREEN if r%2 else WHITE,"left",True,wrap=True); ws.row_dimensions[r].height=32; r+=1
wb.save(MP)
import shutil; shutil.copy(MP,DL+"Космос_Детейлинг_Яндекс_Директ_медиаплан.xlsx")

# ---------- 4. КОММАНДЕР-XLSX (черновик кампании) ----------
cb=openpyxl.Workbook(); ws=cb.active; ws.title="Кампания"; ws.sheet_view.showGridLines=False
cols=["Кампания","Группа","Ключевая фраза","Тип соответствия","Регион","Ставка, ₽","Заголовок 1","Заголовок 2","Текст","Ссылка","Отображаемая ссылка","Быстрые ссылки","Уточнения"]
setw(ws,[24,24,34,14,16,10,34,26,46,30,22,34,40])
for c,h in enumerate(cols,1): C(ws,1,c,h,F(9,True,WHITE),NAVY,"center",True,wrap=True)
ws.row_dimensions[1].height=30
DOMAIN="https://kosmos-detail.ru/"
r=2
for g in GROUP_ORDER:
    rws=sorted(groups[g],key=lambda x:-(x.get("prem_clicks",0) or 0))
    if not rws: continue
    ads=ADS.get(g,[]); land=rws[0]["landing"]
    sl=" | ".join(t for t,_ in SL_COMMON); co=" | ".join(CALLOUTS[:4])
    a0=ads[0] if ads else ("","","")
    for i,x in enumerate(rws):
        # к каждой фразе — первое объявление группы (для остальных вариантов = доп.строки ниже)
        vals=[CAMP,g,f'"{x["kw"]}"',"фразовое","Санкт-Петербург",bid(x),
              a0[0],a0[1],a0[2],DOMAIN+land,f"kosmos-detail.ru/{DISP[g]}",sl,co]
        for c,v in enumerate(vals,1):
            C(ws,r,c,v,F(9),GREEN if (GROUP_ORDER.index(g))%2 else WHITE,"left" if c not in(6,) else "center",True,wrap=True)
        r+=1
    # доп. варианты объявлений группы (без фразы — как доп. объявления группы)
    for a in ads[1:]:
        vals=[CAMP,g,"(доп. объявление группы)","","","",a[0],a[1],a[2],DOMAIN+land,f"kosmos-detail.ru/{DISP[g]}",sl,co]
        for c,v in enumerate(vals,1):
            C(ws,r,c,v,F(9,False,"555555"),GOLDL,"left",True,wrap=True)
        r+=1
ws.freeze_panes="A2"
cb.save(BASE+"Космос_Директ_Коммандер.xlsx"); shutil.copy(BASE+"Космос_Директ_Коммандер.xlsx",DL+"Космос_Директ_Коммандер.xlsx")

print("ADS warnings:",len(warn))
for w in warn: print("  !",w)
print("FILES:")
for p in ["Космос_Детейлинг_Яндекс_Директ_медиаплан.xlsx","Космос_Директ_семантика.csv","Космос_фразы_по_группам.csv","Космос_Директ_Коммандер.xlsx"]:
    import os; print("  ",os.path.exists(BASE+p),p)
print("radius: shows",round(tot_shows*RAD_LO),"-",round(tot_shows*RAD_HI)," clicks",round(tot_rc*RAD_LO),"-",round(tot_rc*RAD_HI))
