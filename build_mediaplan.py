# -*- coding: utf-8 -*-
"""Сборка итогового медиаплана Яндекс.Директ для Космос Детейлинг в один XLSX."""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

DATE = "2026-06-05"
fc = json.load(open("/tmp/forecast.json"))
comp = json.load(open("/tmp/competitors.json"))

# ---------- параметры медиаплана (документируются на листе «Методология») ----------
IS_BASE = 0.50      # доля показов (impression share), базовый сценарий
IS_OPT  = 0.70      # доля показов, активный сценарий
OVERLAP = 0.70      # скидка на пересечение вложенных фраз (антидубль трафика)
EFF_BASE = IS_BASE * OVERLAP     # 0.35
EFF_OPT  = IS_OPT  * OVERLAP     # 0.49
CR = [0.03, 0.05, 0.08]          # конверсия сайт→заявка (допущение, до запуска)
LEAD2SALE = 0.35                 # заявка→клиент (допущение)

GROUP_ORDER = ["Детейлинг общий","Детейлинг · Васильевский остров","Детейлинг-мойка",
"Оклейка / PPF","Полировка кузова","Керамика / защита","Тонировка","Оптика / фары",
"Химчистка и кожа","Предпродажная подготовка"]
GROUP_PHASE = {"Детейлинг · Васильевский остров":1,"Химчистка и кожа":1,"Полировка кузова":1,
"Керамика / защита":1,"Детейлинг общий":1,"Детейлинг-мойка":1,"Предпродажная подготовка":1,
"Оклейка / PPF":2,"Тонировка":2,"Оптика / фары":2}
GROUP_TICKET = {"Детейлинг общий":10000,"Детейлинг · Васильевский остров":12000,"Детейлинг-мойка":4000,
"Оклейка / PPF":90000,"Полировка кузова":12000,"Керамика / защита":20000,"Тонировка":8000,
"Оптика / фары":9000,"Химчистка и кожа":12000,"Предпродажная подготовка":12000}

def temp(r):
    k=r["kw"]
    if any(x in k for x in ["цена","цены","стоимост","недорого","сколько стоит","спб","петербург","василе","рядом","заказать","сделать","записа"]):
        return "ГОР"
    return "ТЕП"

# ---------- расчёт по фразам ----------
for r in fc:
    pc=r.get("prem_clicks",0) or 0
    r["rc_base"]=pc*EFF_BASE
    r["rc_opt"]=pc*EFF_OPT
    cmin=r.get("prem_min",0) or 0; cmax=r.get("prem_max",0) or 0
    r["cpc_mid"]=(cmin+cmax)/2 if (cmin or cmax) else 0
    r["bud_min"]=r["rc_base"]*cmin
    r["bud_max"]=r["rc_base"]*cmax
    r["bud_mid"]=r["rc_base"]*r["cpc_mid"]
    r["temp"]=temp(r)

def gsum(rows,key): return sum(r.get(key,0) or 0 for r in rows)

groups={g:[r for r in fc if r["group"]==g] for g in GROUP_ORDER}

# ---------- стили ----------
NAVY="0B1F3A"; GOLD="C8A24B"; GOLDL="F3E9CE"; DARK="11151F"; GREY="F2F4F7"; LINE="D9DEE7"
WHITE="FFFFFF"; GREEN="E7F6EC"; RED="FBEAEA"; BLUE="EAF1FB"
def F(sz=11,b=False,color="1A1A1A",name="Calibri"): return Font(name=name,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color=LINE)
border=Border(left=thin,right=thin,top=thin,bottom=thin)
def setw(ws,widths):
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
def cell(ws,r,c,v,font=None,fillc=None,al=None,bord=False,fmt=None,wrap=False):
    x=ws.cell(r,c,v)
    if font: x.font=font
    if fillc: x.fill=fill(fillc)
    x.alignment=Alignment(horizontal=al or "left",vertical="center",wrap_text=wrap)
    if bord: x.border=border
    if fmt: x.number_format=fmt
    return x

wb=openpyxl.Workbook()

# =========================================================
# ЛИСТ 1 — СТРАТЕГИЯ И МЕДИАПЛАН
# =========================================================
ws=wb.active; ws.title="1. Стратегия и медиаплан"; ws.sheet_view.showGridLines=False
setw(ws,[3,34,20,18,18,18,16])
ws.merge_cells("B2:G2"); cell(ws,2,2,"КОСМОС ДЕТЕЙЛИНГ — медиаплан Яндекс.Директ",F(18,True,WHITE),NAVY,"left")
ws.row_dimensions[2].height=34
ws.merge_cells("B3:G3"); cell(ws,3,2,f"Детейлинг-студия · Санкт-Петербург, Васильевский остров (ул. Кораблестроителей, 14) · регион показа: СПб (Geo ID 2) · дата расчёта: {DATE}",F(10,False,WHITE),"35506E")
ws.row_dimensions[3].height=22

# Сводка
totals={
 "kw":len(fc),
 "shows":gsum(fc,"d_shows"),
 "prem_clicks":gsum(fc,"prem_clicks"),
 "rc_base":gsum(fc,"rc_base"),
 "rc_opt":gsum(fc,"rc_opt"),
 "bud_min":gsum(fc,"bud_min"),
 "bud_mid":gsum(fc,"bud_mid"),
 "bud_max":gsum(fc,"bud_max"),
}
phase1=[r for r in fc if GROUP_PHASE[r["group"]]==1]
p1_rc=gsum(phase1,"rc_base"); p1_mid=gsum(phase1,"bud_mid"); p1_min=gsum(phase1,"bud_min"); p1_max=gsum(phase1,"bud_max")

row=5
cell(ws,row,2,"КЛЮЧЕВЫЕ ЦИФРЫ (в месяц, прогноз)",F(12,True,NAVY)); row+=1
kp=[
 ("Ключевых фраз в ядре", f"{totals['kw']}", "строго детейлинг + мойка + плёнки/тонировка/оптика + предпродажная"),
 ("Групп объявлений", f"{len([g for g in groups if groups[g]])}", "каждая → своя посадочная страница сайта"),
 ("Прогноз показов/мес (Директ)", f"{round(totals['shows']):,}".replace(',',' '), "суммарно по ядру, регион СПб"),
 ("Реалистичные клики/мес", f"{round(totals['rc_base'])}–{round(totals['rc_opt'])}", "премиум-размещение, доля показов 50–70%, с учётом пересечений"),
 ("Бюджет/мес (вся семантика)", f"{round(totals['bud_min']):,}".replace(',',' ')+f" – {round(totals['bud_max']):,}".replace(',',' ')+" ₽", "клики × цена клика (мин–макс по прогнозу)"),
 ("Рекоменд. старт (Фаза 1)", f"{round(p1_min):,}".replace(',',' ')+f" – {round(p1_max):,}".replace(',',' ')+" ₽/мес", "локальные + маржинальные группы, без дорогих плёнок"),
]
for label,val,note in kp:
    cell(ws,row,2,label,F(10,True),GREY,bord=True)
    ws.merge_cells(start_row=row,start_column=3,end_row=row,end_column=4)
    cell(ws,row,3,val,F(11,True,NAVY),GOLDL,"left",True)
    ws.merge_cells(start_row=row,start_column=5,end_row=row,end_column=7)
    cell(ws,row,5,note,F(9,False,"555555"),WHITE,"left",True,wrap=True); row+=1
row+=1

# Сценарии бюджета
cell(ws,row,2,"СЦЕНАРИИ БЮДЖЕТА (в месяц)",F(12,True,NAVY)); row+=1
hdr=["Сценарий","Охват семантики","Доля показов (IS)","Клики/мес","Цена клика","Бюджет/мес, ₽"]
for c,h in enumerate(hdr,2): cell(ws,row,c,h,F(10,True,WHITE),NAVY,"center",True)
row+=1
scen=[
 ("A. Старт / тест","Фаза 1 (локально+маржа)","50%",f"{round(p1_rc)}","средняя",f"{round(p1_mid):,}".replace(',',' ')),
 ("B. Базовый","Вся семантика","50%",f"{round(totals['rc_base'])}","средняя",f"{round(totals['bud_mid']):,}".replace(',',' ')),
 ("C. Максимум охвата","Вся семантика","70%",f"{round(totals['rc_opt'])}","ближе к макс.",f"{round(totals['bud_max']):,}".replace(',',' ')),
]
for s in scen:
    for c,v in enumerate(s,2):
        cell(ws,row,c,v,F(10,c==2),WHITE if c!=2 else GOLDL,"center" if c>2 else "left",True)
    row+=1
row+=1

# Воронка / лиды
cell(ws,row,2,"ПРОГНОЗ ЗАЯВОК И CPL (сценарий B «Базовый», допущения по конверсии)",F(12,True,NAVY)); row+=1
hdr=["Конверсия сайт→заявка","Заявок/мес","Бюджет/мес, ₽","CPL (цена заявки), ₽","Клиентов/мес*","CAC (цена клиента), ₽"]
for c,h in enumerate(hdr,2): cell(ws,row,c,h,F(10,True,WHITE),NAVY,"center",True)
row+=1
clicks_b=totals['rc_base']; bud_b=totals['bud_mid']
for cr in CR:
    leads=clicks_b*cr; cpl=bud_b/leads if leads else 0
    clients=leads*LEAD2SALE; cac=bud_b/clients if clients else 0
    vals=[f"{int(cr*100)}%",f"{round(leads)}",f"{round(bud_b):,}".replace(',',' '),f"{round(cpl):,}".replace(',',' '),f"{round(clients)}",f"{round(cac):,}".replace(',',' ')]
    for c,v in enumerate(vals,2): cell(ws,row,c,v,F(10,c==4 or c==6),GREEN if cr==0.05 else WHITE,"center",True)
    row+=1
cell(ws,row,2,"* заявка→клиент принята за 35% (допущение). Средний чек по нишам: полировка ~12 000 ₽, керамика ~20 000 ₽, химчистка ~12 000 ₽, оклейка/PPF 70 000 ₽+. Даже при CPL 1 500–3 000 ₽ услуги маржинальны.",F(8,False,"777777"),wrap=True)
ws.merge_cells(start_row=row,start_column=2,end_row=row,end_column=7); ws.row_dimensions[row].height=28; row+=2

# Рекомендации/стратегия
cell(ws,row,2,"СТРАТЕГИЯ ЗАПУСКА",F(12,True,NAVY)); row+=1
recs=[
 "1. Гео: РФ→СПб (Geo ID 2) + радиус 5–7 км вокруг Кораблестроителей 14 (Васильевский остров, Намыв, остров Декабристов). Гиперлокальные ключи («василеостровский», «рядом») держим на минимальных ставках — высокий интент, почти нулевая частота.",
 "2. Фаза 1 (1–2 мес, тест): локальные + маржинальные группы — Детейлинг+гео, Химчистка, Полировка, Керамика, Предпродажная, Детейлинг-мойка. Бюджет см. «Сценарий A». Цель — собрать реальную конверсию и CPL.",
 "3. Фаза 2 (масштаб): добавить дорогие группы — Оклейка/PPF (CPC ~650 ₽!), Тонировка, Оптика/фары. Включать после подтверждённого ROI и накопленных кейсов до/после.",
 "4. Структура: 1 группа объявлений = 1 кластер = 1 посадочная страница сайта (см. колонку «Посадочная»). Это повышает Quality Score и снижает цену клика.",
 "5. Тип соответствия: фразовое (кавычки), на старте — с операторами; брендовые запросы конкурентов вынести в отдельную кампанию-перехват (осторожно) либо в минус-слова.",
 "6. Объявления: бить в боли (фикс. смета, фото/видео-отчёт, гарантия, отдельные боксы — не мойка). Быстрые ссылки на услуги, уточнения, цены «от», адрес и кнопки WhatsApp/Telegram.",
 "7. ОБЯЗАТЕЛЬНО параллельно: карточки в Яндекс.Картах и 2ГИС — SEO-выдача занята агрегаторами (Я.Карты, 2ГИС, Avito, Zoon), органика для новой студии медленная. Карты дают локальный трафик сразу.",
 "8. Аналитика: Яндекс.Метрика + цели (заявка, клик по WhatsApp/тел.), коллтрекинг. Без целей оптимизация невозможна. Через 2–4 недели — чистка по поисковым запросам и ставкам.",
]
for t in recs:
    ws.merge_cells(start_row=row,start_column=2,end_row=row,end_column=7)
    cell(ws,row,2,t,F(10,False,"222222"),WHITE if row%2 else GREY,"left",True,wrap=True)
    ws.row_dimensions[row].height=30; row+=1
ws.freeze_panes="A5"

# =========================================================
# ЛИСТ 2 — СЕМАНТИКА + ПРОГНОЗ
# =========================================================
ws=wb.create_sheet("2. Семантика + прогноз"); ws.sheet_view.showGridLines=False
cols=["№","Ключевая фраза","Группа объявлений","Посадочная страница","Тип","Спрос Wordstat СПб/мес","Прогноз показов (Директ)","Клики премиум (100% IS)","Реал. клики/мес (IS50%×0.7)","CPC мин, ₽","CPC макс, ₽","Бюджет/мес мин, ₽","Бюджет/мес макс, ₽","Темп-ра"]
setw(ws,[5,40,26,24,8,16,16,15,16,11,11,15,15,9])
ws.merge_cells("A1:N1"); cell(ws,1,1,"СЕМАНТИЧЕСКОЕ ЯДРО + ПРОГНОЗ ЯНДЕКС.ДИРЕКТ (регион СПб, премиум-размещение)",F(13,True,WHITE),NAVY)
ws.row_dimensions[1].height=26
for c,h in enumerate(cols,1): cell(ws,2,c,h,F(9,True,WHITE),NAVY,"center",True,wrap=True)
ws.row_dimensions[2].height=42
r=3; n=0
INT="#,##0"
for g in GROUP_ORDER:
    rows=sorted(groups[g],key=lambda x:-(x.get("rc_base",0)))
    if not rows: continue
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=14)
    cell(ws,r,1,f"▸ {g}  ·  {rows[0]['landing']}  ·  фраз: {len(rows)}  ·  Фаза {GROUP_PHASE[g]}",F(10,True,DARK),GOLDL); r+=1
    for x in rows:
        n+=1
        vals=[n,x["kw"],x["group"],x["landing"],"Фраз.",x.get("wordstat",0),round(x.get("d_shows",0)),
              round(x.get("prem_clicks",0)),round(x.get("rc_base",0),1),
              round(x.get("prem_min",0)),round(x.get("prem_max",0)),
              round(x.get("bud_min",0)),round(x.get("bud_max",0)),x["temp"]]
        for c,v in enumerate(vals,1):
            al="center" if c in (1,5,6,7,8,9,10,11,12,13,14) else "left"
            fc_=WHITE
            if c==14: fc_=RED if v=="ГОР" else BLUE
            cell(ws,r,c,v,F(9, c==2),fc_,al,True,fmt=INT if c in(6,7,8,10,11,12,13) else None)
        r+=1
# итоговая строка
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5); cell(ws,r,1,"ИТОГО по ядру",F(10,True,WHITE),NAVY,"right")
tot=[totals["shows"]>0 and gsum(fc,"wordstat") or 0,round(totals["shows"]),round(totals["prem_clicks"]),round(totals["rc_base"],1),"","",round(totals["bud_min"]),round(totals["bud_max"]),""]
cell(ws,r,6,gsum(fc,"wordstat"),F(10,True,WHITE),NAVY,"center",fmt=INT)
cell(ws,r,7,round(totals["shows"]),F(10,True,WHITE),NAVY,"center",fmt=INT)
cell(ws,r,8,round(totals["prem_clicks"]),F(10,True,WHITE),NAVY,"center",fmt=INT)
cell(ws,r,9,round(totals["rc_base"]),F(10,True,WHITE),NAVY,"center",fmt=INT)
cell(ws,r,10,"",F(10),NAVY); cell(ws,r,11,"",F(10),NAVY)
cell(ws,r,12,round(totals["bud_min"]),F(10,True,WHITE),NAVY,"center",fmt=INT)
cell(ws,r,13,round(totals["bud_max"]),F(10,True,WHITE),NAVY,"center",fmt=INT)
cell(ws,r,14,"",F(10),NAVY)
ws.freeze_panes="A3"; ws.auto_filter.ref=f"A2:N2"

# =========================================================
# ЛИСТ 3 — МЕДИАПЛАН ПО ГРУППАМ
# =========================================================
ws=wb.create_sheet("3. Медиаплан по группам"); ws.sheet_view.showGridLines=False
cols=["Группа объявлений","Посадочная","Фаза","Ключей","Показов/мес","Реал. клики/мес","CPC ср., ₽","Бюджет/мес мин, ₽","Бюджет/мес макс, ₽","Заявок/мес (5%)","CPL, ₽","Ср. чек, ₽","Приоритет"]
setw(ws,[28,22,7,8,13,15,11,15,15,14,11,11,24])
ws.merge_cells("A1:M1"); cell(ws,1,1,"МЕДИАПЛАН ПО ГРУППАМ (в месяц, сценарий B, IS 50%)",F(13,True,WHITE),NAVY); ws.row_dimensions[1].height=26
for c,h in enumerate(cols,1): cell(ws,2,c,h,F(9,True,WHITE),NAVY,"center",True,wrap=True)
ws.row_dimensions[2].height=40
r=3
def prio(g):
    return {1:"Высокий — старт",2:"Средний — масштаб"}[GROUP_PHASE[g]]
for g in GROUP_ORDER:
    rows=groups[g]
    if not rows: continue
    rc=gsum(rows,"rc_base"); bmin=gsum(rows,"bud_min"); bmax=gsum(rows,"bud_max")
    cpcs=[r2["cpc_mid"] for r2 in rows if r2["cpc_mid"]]; cpc=sum(cpcs)/len(cpcs) if cpcs else 0
    leads=rc*0.05; cpl=((bmin+bmax)/2)/leads if leads else 0
    vals=[g,rows[0]["landing"],f"Фаза {GROUP_PHASE[g]}",len(rows),round(gsum(rows,"d_shows")),round(rc),round(cpc),round(bmin),round(bmax),round(leads,1),round(cpl),GROUP_TICKET[g],prio(g)]
    fillc=GREEN if GROUP_PHASE[g]==1 else BLUE
    for c,v in enumerate(vals,1):
        cell(ws,r,c,v,F(9,c==1),fillc,"left" if c in(1,2,13) else "center",True,fmt=INT if c in(5,6,7,8,9,11,12) else None)
    r+=1
# total
tt=[("ИТОГО","",f"",len(fc),round(totals["shows"]),round(totals["rc_base"]),"",round(totals["bud_min"]),round(totals["bud_max"]),round(totals["rc_base"]*0.05),round(totals["bud_mid"]/(totals["rc_base"]*0.05)) if totals["rc_base"] else 0,"","")]
for v in tt:
    for c,val in enumerate(v,1):
        cell(ws,r,c,val,F(10,True,WHITE),NAVY,"left" if c==1 else "center",True,fmt=INT if c in(5,6,8,9,10,11) else None)
ws.freeze_panes="A3"

# =========================================================
# ЛИСТ 4 — КОНКУРЕНТЫ
# =========================================================
ws=wb.create_sheet("4. Конкуренты"); ws.sheet_view.showGridLines=False
setw(ws,[4,30,14,46])
ws.merge_cells("B1:E1"); cell(ws,1,2,"КОНКУРЕНТЫ В ЯНДЕКСЕ: SEO-выдача и реклама (регион СПб)",F(13,True,WHITE),NAVY); ws.row_dimensions[1].height=26
r=3
cell(ws,r,2,"1. КТО В ТОП-10 ОРГАНИКИ (частота попадания по 10 ключам ядра)",F(11,True,NAVY)); r+=1
for c,h in enumerate(["Домен","Появлений","Тип игрока"],2): cell(ws,r,c,h,F(10,True,WHITE),NAVY,"center",True)
r+=1
def kind(d):
    agg=["yandex.ru","uslugi.yandex.ru","2gis.ru","avito.ru","zoon.ru","yell.ru","orgpage.ru","vc.ru"]
    if d in agg: return "Агрегатор / каталог"
    return "Детейлинг-студия / сервис"
for d,c in comp["counts"][:22]:
    fillc=GOLDL if kind(d).startswith("Агрег") else WHITE
    cell(ws,r,2,d,F(10,kind(d)=="Детейлинг-студия / сервис"),fillc,"left",True)
    cell(ws,r,3,c,F(10),fillc,"center",True)
    cell(ws,r,4,kind(d),F(9,False,"555555"),fillc,"left",True); r+=1
r+=1
cell(ws,r,2,"2. ВЫВОД ПО SEO",F(11,True,NAVY)); r+=1
seo=[
 "Топ выдачи занят АГРЕГАТОРАМИ: Яндекс.Карты/Услуги, 2ГИС, Avito, Zoon, подборки vc.ru «ТОП-10 студий». Пробиться в органику новой студии быстро не выйдет.",
 "Из студий в топе: StarDust, spbdetailing.ru, chameleon-spb, platinum-garage, m3-spb, detailing-stars, acpolimer, ac78, 1st-garage, bastion-center, solartek, autoestetika.",
 "Стратегия SEO: (1) карточки в Яндекс.Картах и 2ГИС с отзывами — приоритет №1; (2) попадание в подборки-агрегаторы; (3) гео-långtail страницы (детейлинг Васильевский/Намыв), где агрегаторы слабее.",
]
for t in seo:
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4); cell(ws,r,2,t,F(10,False,"222222"),GREY if r%2 else WHITE,"left",True,wrap=True); ws.row_dimensions[r].height=30; r+=1
r+=1
cell(ws,r,2,"3. КОНКУРЕНТЫ В РЕКЛАМЕ (Директ) и на локальном рынке В.О.",F(11,True,NAVY)); r+=1
for c,h in enumerate(["Конкурент / сайт","Профиль","Чем сильны / на что смотреть"],2): cell(ws,r,c,h,F(10,True,WHITE),NAVY,"center",True)
r+=1
market=[
 ("Quality Detailing — qdet.ru","В.О., 6-я линия","Премиум, гарантия, видео-кейсы, школа. Сильный локальный ориентир."),
 ("StarDust — stardustspb.ru","В.О., 27-я линия","Прозрачные пакеты Light/Optimal/Pro, цены, ~180 отзывов."),
 ("Grade Restyling — graderc.ru","В.О., Морская Слава","Премиум-визуал, сложные работы; цены закрыты (запрос)."),
 ("Performance — performance-spb.ru","СПб","Фокус на китайские авто; реклама/Директ."),
 ("Detailing Piter — detailing-piter.ru","ТРК Питерлэнд","Сильное SEO, бронирование стёкол."),
 ("Bulldog / Parking — Кораблестроителей 14","ваш адрес","Прямые соседи: поток мойки Parking. Перехват в премиум-формат."),
 ("Агрегаторы: Я.Карты, 2ГИС, Avito, Zoon","площадки","Забирают и рекламу, и органику. Нужны платные карточки/размещения."),
]
for nm,pr,note in market:
    cell(ws,r,2,nm,F(10,True),WHITE,"left",True); cell(ws,r,3,pr,F(9),WHITE,"left",True); cell(ws,r,4,note,F(9,False,"444444"),WHITE,"left",True,wrap=True); ws.row_dimensions[r].height=28; r+=1
r+=1
cell(ws,r,2,"4. КАК ОБЫГРАТЬ",F(11,True,NAVY)); r+=1
win=[
 "Ставка на ДОВЕРИЕ, а не «премиум на словах»: фикс. смета до работ, фото/видео-отчёт, договор/акт, гарантия, отдельные чистые боксы (не поток мойки) — этого нет у большинства.",
 "Локальное удобство: «рядом на Васильевском, без очередей, запись у дома» — против студий в центре/промзоне.",
 "Дорогие плёнки (CPC ~650 ₽) — не лить на старте; сначала маржинальные химчистка/полировка/керамика с понятным CPL.",
]
for t in win:
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4); cell(ws,r,2,t,F(10,False,"222222"),GREEN if r%2 else WHITE,"left",True,wrap=True); ws.row_dimensions[r].height=30; r+=1

# =========================================================
# ЛИСТ 5 — МИНУС-СЛОВА
# =========================================================
ws=wb.create_sheet("5. Минус-слова"); ws.sheet_view.showGridLines=False
setw(ws,[4,30,64])
ws.merge_cells("B1:C1"); cell(ws,1,2,"МИНУС-СЛОВА (единый список на кампанию)",F(13,True,WHITE),NAVY); ws.row_dimensions[1].height=26
blocks=[
 ("Самостоятельность / DIY","своими руками, в домашних условиях, самостоятельно, как сделать, как клеить, как нанести, содой, лимоном, паста, полироль, средство, набор, состав, рецепт"),
 ("Инфо / обучение","что такое, что это, видео, обучение, курсы, курс, школа, мастер класс, реферат, отзывы (если не нужны), форум, вики"),
 ("Работа","вакансии, вакансия, работа, зарплата, требуется, ученик, мастер требуется, резюме"),
 ("Товары / опт","купить, магазин, цена пленки, пленка купить, оптом, бу, б у, авито (как площадка), озон, валберис, wildberries, набор для, оборудование, аппарат, станок, инструмент, пылесос"),
 ("Юр./ГИБДД (для тонировки)","гост, штраф, разрешенная, разрешено, можно ли, процент, светопропускание, нормы, пдд, наказание, растонировка, размотать, снять тонировку"),
 ("Не наш профиль (услуги вне ядра)","винил, цветная, матовая (как смена цвета), антихром, диски, покраска дисков, шумоизоляция, шумка, доводчики, сигнализация, сигнализации, перетяжка, перешив, вмятины, вмятин, pdr, антидождь (как отдельная услуга), грузовик, фура, мотоцикл, такси"),
 ("Другие города","москва, краснодар, сочи, екатеринбург, новосибирск, казань, нижний новгород, ростов, самара, уфа, пермь, воронеж, калининград, мурманск + остальные регионы РФ"),
 ("Другие районы СПб (опц., если узкое гео)","приморский, выборгский, московский, фрунзенский, невский, купчино, колпино, пушкин, кронштадт, всеволожск, мурино, кудрово, парнас"),
 ("Бренды конкурентов (в минус ИЛИ отд. кампания-перехват)","stardust, quality detailing, qdet, grade, бульдог, bulldog, малибу, гма, аврора, тритон, саппо, detailing piter, performance"),
 ("Бесплатно / прочее","бесплатно, дешево (опц.), скачать, онлайн калькулятор пленки, франшиза, бизнес, оквэд"),
]
r=3
for title,words in blocks:
    cell(ws,r,2,title,F(10,True,WHITE),GOLD,"left",True)
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=3)
    cell(ws,r,3,words,F(10,False,"222222"),GREY if r%2 else WHITE,"left",True,wrap=True)
    ws.row_dimensions[r].height=max(18, 16*(len(words)//70+1)); r+=1
cell(ws,r+1,2,"Совет:",F(10,True,NAVY)); ws.merge_cells(start_row=r+1,start_column=3,end_row=r+1,end_column=3)
cell(ws,r+1,3,"Минус-слова уровня кампании задать до старта; через 1–2 недели чистить по «Поисковым запросам» в Метрике/Директе и пополнять список.",F(10,False,"444444"),wrap=True)

# =========================================================
# ЛИСТ 6 — МЕТОДОЛОГИЯ
# =========================================================
ws=wb.create_sheet("6. Методология и источники"); ws.sheet_view.showGridLines=False
setw(ws,[4,34,70])
ws.merge_cells("B1:C1"); cell(ws,1,2,"МЕТОДОЛОГИЯ, ДОПУЩЕНИЯ И ИСТОЧНИКИ",F(13,True,WHITE),NAVY); ws.row_dimensions[1].height=26
meta=[
 ("Дата расчёта",DATE),
 ("Регион",f"Санкт-Петербург (Yandex Geo ID 2). Рекомендуемое гео в кампании — СПб + радиус 5–7 км вокруг Кораблестроителей 14."),
 ("Источник спроса (Wordstat)","XMLRiver Wordstat API, параметр regions=2 (СПб). Колонка «Спрос Wordstat» = частота показов фразы/мес по региону. Это спрос, НЕ клики."),
 ("Источник прогноза (Директ)","API Яндекс.Директ, метод CreateNewForecast/GetForecast (Live v4), GeoID=[2], валюта RUB. Колонки Показы/Клики премиум/CPC взяты из прогноза."),
 ("Цена клика (CPC)","Диапазон премиум-размещения (PremiumMin–PremiumMax) — спецразмещение вверху поиска, где сосредоточены клики. Цены прогноза включают НДС."),
 ("Доля показов (IS)","Реальные клики = клики премиум × IS. Базовый сценарий IS=50%, активный IS=70%. 100% показов недостижимо — поэтому хейркат обязателен."),
 ("Пересечение фраз (overlap)","Вложенные фразы («полировка кузова» ⊃ «полировка кузова спб») делят один и тот же спрос. Применён коэффициент 0.70, чтобы не задваивать трафик и бюджет."),
 ("Итоговый множитель кликов","IS × overlap = 0.35 (базовый) и 0.49 (активный). Реал. клики = клики премиум × множитель."),
 ("Бюджет","Бюджет фразы = реал. клики × CPC. Диапазон мин–макс = по CPCmin и CPCmax. Суммирование по группам корректно за счёт overlap."),
 ("Конверсия в заявку","ДОПУЩЕНИЕ (до запуска реальных данных нет): сайт→заявка 3% / 5% / 8%. Заявка→клиент — 35%. После старта заменить фактическими целями Метрики."),
 ("Средний чек","Ориентиры рынка В.О. (из исследования): полировка 8–16к, керамика 12–30к, химчистка 10–15к, оклейка/PPF от 70к. Использованы для оценки окупаемости/CAC."),
 ("Конкуренты","SEO — топ-10 органики Яндекса (XMLRiver Yandex SERP, loc=2) по 10 ключам ядра. Реклама — Яндекс XML не отдаёт платные блоки; рекламные конкуренты оценены по уровню CPC и рыночным данным."),
 ("Тип соответствия","Весь прогноз и список — под фразовое соответствие (единообразно, чтобы CPC/клики были сопоставимы)."),
 ("Гиперлокальные ключи","«василеостровский», «намыв», «рядом» имеют ~0 частоты Wordstat, но высокий интент. Берём на минимальных ставках; в объёме трафика не учитываем."),
 ("ВАЖНО — статус цифр","Это ПЛАН/ПРОГНОЗ, а не факт. Реальные ставки и клики зависят от качества объявлений, посадочных, конкуренции и времени. Прогноз Директа — ориентир, цифры округлены."),
]
r=3
for k,v in meta:
    cell(ws,r,2,k,F(10,True,NAVY),GREY,"left",True)
    cell(ws,r,3,v,F(10,False,"222222"),WHITE,"left",True,wrap=True)
    ws.row_dimensions[r].height=max(20,15*(len(v)//72+1)); r+=1

out1="/Users/zif91/Programs_py/КОсмос/Космос_Детейлинг_Яндекс_Директ_медиаплан.xlsx"
out2="/Users/zif91/Downloads/Космос_Детейлинг_Яндекс_Директ_медиаплан.xlsx"
wb.save(out1)
import shutil; shutil.copy(out1,out2)
print("SAVED:",out1)
print("COPY :",out2)
print("totals: kw",totals["kw"]," shows",round(totals["shows"])," prem_clicks",round(totals["prem_clicks"]),
      " rc_base",round(totals["rc_base"])," budget_mid",round(totals["bud_mid"]),
      " budget_range",round(totals["bud_min"]),"-",round(totals["bud_max"]))
print("phase1 budget mid:",round(p1_mid)," range",round(p1_min),"-",round(p1_max))
